import openpyxl
from pathlib import Path


def parse_xlsx(file_path: str) -> list[dict]:
    """
    Extract text from all sheets using header-aware formatting.
    Each data row becomes "Header: value" pairs — more semantic than pipe-delimited.
    Returns list of {text, metadata} dicts.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # First non-empty row is treated as the header
        headers = [str(cell).strip() if cell is not None else f"Col{i}"
                   for i, cell in enumerate(rows[0])]

        formatted_rows = []
        for row in rows[1:]:
            # Skip entirely empty rows
            if all(cell is None for cell in row):
                continue
            pairs = []
            for header, cell in zip(headers, row):
                if cell is not None and str(cell).strip():
                    pairs.append(f"{header}: {str(cell).strip()}")
            if pairs:
                formatted_rows.append(" | ".join(pairs))

        if not formatted_rows:
            continue

        # Prepend sheet name as context
        text = f"[Sheet: {sheet_name}]\n" + "\n".join(formatted_rows)

        sections.append({
            "text": text,
            "metadata": {
                "source": Path(file_path).name,
                "file_path": str(file_path),
                "file_type": "xlsx",
                "sheet": sheet_name,
                "page": 1
            }
        })

    wb.close()
    return sections
